from __future__ import annotations
import math, re, json, os
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import sys
sys.path.append('/mnt/data')
from utils import IV, Butterfly, Regime_Label

WORKBOOK = Path('Option_Data.xlsx')
OUTDIR = Path('Hedging_Expiriment/')
OUTDIR.mkdir(exist_ok=True)

# Sample of expiries; each gives one trade date = expiry - 1 business day.
EXPIRIES = [
    '2026-01-29','2026-01-30','2026-02-02','2026-02-03',
    '2026-02-04','2026-02-05','2026-02-06','2026-02-09'
]

H = 0.01
TOL = 0.0025
MIN_PRICE = 0.05
MAX_ABS_K = 0.03
TARGET_ABS_K = 0.012
DT_YEAR = 1.0/(252.0*390.0)
R = 0.0


def parse_contract_name(name: str):
    if not isinstance(name, str):
        return None
    m = re.search(r"(\d{2}/\d{2}/\d{2})\s+([CP])\s*([0-9]+(?:\.[0-9]+)?)", name)
    if not m:
        m = re.search(r"(\d{2}/\d{2}/\d{2})\s+([CP])([0-9]+(?:\.[0-9]+)?)", name)
    if not m:
        return None
    exp = pd.to_datetime(m.group(1), format="%m/%d/%y").normalize()
    return {"expiry_date": exp, "cp": m.group(2), "K": float(m.group(3))}


def parse_sheet(ws, sheet_name):
    rows = ws.iter_rows(values_only=True)
    try:
        row0 = next(rows)
        row1 = next(rows)
    except StopIteration:
        return pd.DataFrame()
    blocks = []
    for i in range(0, len(row0), 5):
        if i + 4 >= len(row0):
            break
        info = parse_contract_name(row0[i])
        if info is None:
            continue
        labels = [str(x).strip().lower() if x is not None else '' for x in row1[i:i+5]]
        if labels != ['dates','open','close','value','volume']:
            continue
        blocks.append((i, info))
    out = []
    for row in rows:
        for i, info in blocks:
            ts = row[i]
            close = row[i+2]
            if ts is None or close is None:
                continue
            try:
                if isinstance(ts, (int, float, np.integer, np.floating)) and np.isfinite(ts) and ts > 10000:
                    ts = pd.to_datetime(float(ts), unit='D', origin='1899-12-30')
                else:
                    ts = pd.to_datetime(ts)
                mid = float(close)
            except Exception:
                continue
            vol = row[i+4]
            try:
                vol = float(vol) if vol is not None else 0.0
            except Exception:
                vol = 0.0
            out.append((ts, mid, vol, info['expiry_date'], info['cp'], info['K']))
    if not out:
        return pd.DataFrame()
    df = pd.DataFrame(out, columns=['timestamp','mid','volume','expiry_date','cp','K'])
    df['trade_date'] = df['timestamp'].dt.normalize()
    return df


def robust_spot_from_pairs(quotes: pd.DataFrame) -> pd.DataFrame:
    q = quotes.copy()
    q['tbin'] = q['timestamp'].dt.floor('10s')
    piv_px = q.pivot_table(index=['tbin','K'], columns='cp', values='mid', aggfunc='last').reset_index()
    piv_vol = q.pivot_table(index=['tbin','K'], columns='cp', values='volume', aggfunc='sum').reset_index()
    piv_px.columns.name = None
    piv_vol.columns.name = None
    if 'C' not in piv_px.columns or 'P' not in piv_px.columns:
        return pd.DataFrame(columns=['minute','S'])
    pairs = piv_px.merge(piv_vol.rename(columns={'C':'vol_c','P':'vol_p'}), on=['tbin','K'], how='left')
    pairs = pairs.dropna(subset=['C','P']).copy()
    if pairs.empty:
        return pd.DataFrame(columns=['minute','S'])
    pairs['pair_volume'] = pairs[['vol_c','vol_p']].fillna(0.0).sum(axis=1)
    pairs['S_parity'] = pairs['C'] - pairs['P'] + pairs['K']
    pairs = pairs[np.isfinite(pairs['S_parity']) & (pairs['S_parity'] > 0)].copy()
    if pairs.empty:
        return pd.DataFrame(columns=['minute','S'])
    rows=[]
    for ts, g in pairs.groupby('tbin'):
        x = g['S_parity'].to_numpy(float)
        w = np.sqrt(np.clip(g['pair_volume'].to_numpy(float), 1.0, None))
        med = np.nanmedian(x)
        mad = np.nanmedian(np.abs(x-med))
        if np.isfinite(mad) and mad > 0:
            keep = np.abs(x-med) <= 4.0*1.4826*mad
            x = x[keep]; w = w[keep]
        if len(x) == 0:
            continue
        rows.append((pd.Timestamp(ts).floor('1min'), float(np.sum(w*x)/np.sum(w))))
    spot = pd.DataFrame(rows, columns=['minute','S']).sort_values('minute')
    spot = spot.groupby('minute', as_index=False)['S'].last()
    return spot


def iv_for_quotes(quotes: pd.DataFrame, spot: pd.DataFrame) -> pd.DataFrame:
    q = quotes.copy()
    q['minute'] = q['timestamp'].dt.floor('1min')
    q = q.sort_values('timestamp').groupby(['minute','cp','K','expiry_date'], as_index=False).tail(1).copy()
    q = q.merge(spot, on='minute', how='left')
    q['S'] = q['S'].ffill().bfill()
    q['expiry_dt'] = q['expiry_date'] + pd.Timedelta(hours=16)
    q['tau'] = (q['expiry_dt'] - q['minute']).dt.total_seconds() / (365.0*24.0*3600.0)
    q = q[(q['tau'] > 0) & (q['mid'] >= MIN_PRICE) & (q['S'] > 0)].copy()
    q['k'] = np.log(q['K']/q['S'])
    q = q[np.abs(q['k']) <= MAX_ABS_K].copy()
    if q.empty:
        return q
    ivs=[]; vegas=[]
    for price, S, K, tau, cp in q[['mid','S','K','tau','cp']].itertuples(index=False, name=None):
        iv = IV.implied_vol_mid(float(price), float(S), float(K), R, float(tau), str(cp))
        if not np.isfinite(iv):
            ivs.append(np.nan); vegas.append(np.nan); continue
        ivs.append(iv)
        vegas.append(IV.bs_vega(float(S), float(K), R, float(tau), float(iv)))
    q['iv'] = ivs
    q['vega'] = vegas
    q = q[np.isfinite(q['iv']) & np.isfinite(q['vega']) & (q['vega'] > 0)].copy()
    q['trade_date'] = q['minute'].dt.normalize()
    return q[['minute','trade_date','cp','K','mid','S','tau','k','iv','vega','expiry_date']].rename(columns={'minute':'timestamp'})


def local_sigma(x, y, target):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    idx = np.where(np.abs(x-target) <= TOL)[0]
    if len(idx) < 3:
        idx = np.argsort(np.abs(x-target))[:3]
    return float(np.median(y[idx])) if len(idx) else np.nan


def build_features(qiv: pd.DataFrame) -> pd.DataFrame:
    rows=[]
    for ts, g in qiv.groupby('timestamp'):
        x = g['k'].to_numpy(float)
        y = g['iv'].to_numpy(float)
        if len(g) < 6:
            continue
        sL = local_sigma(x,y,-H)
        s0 = local_sigma(x,y,0.0)
        sR = local_sigma(x,y,H)
        if not (np.isfinite(sL) and np.isfinite(s0) and np.isfinite(sR)):
            continue
        skew = (sR - sL)/(2*H)
        curv = 2*(0.5*(sL+sR)-s0)/(H*H)
        rows.append((ts, pd.Timestamp(ts).normalize(), float(np.nanmedian(g['S'])), s0, skew, curv))
    feat = pd.DataFrame(rows, columns=['timestamp','trade_date','S','sigma_atm','skew','curvature_bfly'])
    if feat.empty:
        return feat
    feat = feat.sort_values('timestamp').reset_index(drop=True)
    for col in ['sigma_atm','skew','curvature_bfly']:
        sm = feat.groupby('trade_date')[col].transform(lambda s: s.ewm(span=Butterfly.EWM_SPAN, adjust=False).mean())
        d = sm.groupby(feat['trade_date']).diff().fillna(0.0)
        z = d.groupby(feat['trade_date']).transform(Butterfly.zscore_day)
        feat[f'z_{col}'] = z
        cs = Butterfly.cusum_series(z)
        feat[f'event_{col}'] = cs['event'].values
    feat['event_any'] = feat[[f'event_{c}' for c in ['sigma_atm','skew','curvature_bfly']]].max(axis=1)
    feat['event_score'] = feat[[f'z_{c}' for c in ['sigma_atm','skew','curvature_bfly']]].abs().sum(axis=1)
    feat = Regime_Label.build_regime_labels(feat)
    return feat


def bs_delta(cp, S, K, tau, sigma):
    S=max(float(S),1e-12); K=max(float(K),1e-12); tau=max(float(tau),1e-12); sigma=min(max(float(sigma),1e-8),5.0)
    d1=(math.log(S/K)+(0.5*sigma*sigma)*tau)/(sigma*math.sqrt(tau))
    if str(cp).upper().startswith('C'):
        return 0.5*(1.0+math.erf(d1/math.sqrt(2)))
    return -0.5*(1.0+math.erf(-d1/math.sqrt(2)))


def bs_price(cp, S, K, tau, sigma):
    S=max(float(S),1e-12); K=max(float(K),1e-12); tau=max(float(tau),1e-12); sigma=min(max(float(sigma),1e-8),5.0)
    d1=(math.log(S/K)+(0.5*sigma*sigma)*tau)/(sigma*math.sqrt(tau))
    d2=d1-sigma*math.sqrt(tau)
    ncdf=lambda x: 0.5*(1+math.erf(x/math.sqrt(2)))
    if str(cp).upper().startswith('C'):
        return S*ncdf(d1)-K*ncdf(d2)
    return K*ncdf(-d2)-S*ncdf(-d1)


def smile_iv(S,K,sigma_atm,skew,curvature):
    k=math.log(max(K,1e-12)/max(S,1e-12))
    return min(max(float(sigma_atm)+float(skew)*k+0.5*float(curvature)*k*k,1e-4),3.0)


def surface_delta(cp,S,K,tau,sigma_atm,skew,curvature):
    h=max(1e-3, abs(float(S))*1e-4)
    up=bs_price(cp,S+h,K,tau,smile_iv(S+h,K,sigma_atm,skew,curvature))
    dn=bs_price(cp,max(S-h,1e-8),K,tau,smile_iv(max(S-h,1e-8),K,sigma_atm,skew,curvature))
    return (up-dn)/(2*h)


def shock_surface_delta(cp,S,K,tau,sigma_atm,skew,curvature,event_score,regime):
    boost = 1.0 + (0.18*min(float(event_score), 8.0)/8.0 if int(regime)==1 else 0.0)
    return surface_delta(cp,S,K,tau,min(max(sigma_atm*boost,1e-4),3.0),skew,curvature)


def pick_main_contracts(qiv: pd.DataFrame) -> pd.DataFrame:
    # choose one ATM-ish contract per minute with best abs(k), preferring call for k>=0 and put for k<0 not enforced
    q = qiv.copy()
    q['score'] = np.abs(np.abs(q['k']) - TARGET_ABS_K) + 0.2/(q['vega']+1e-6)
    # Keep one contract_id per day for stability: best coverage near target
    q['contract_id'] = q['cp'].astype(str) + '_' + q['K'].round(6).astype(str)
    cands = q.groupby(['trade_date','contract_id','cp','K'], as_index=False).agg(coverage=('timestamp','size'), mean_abs_k=('k', lambda s: float(np.nanmean(np.abs(s)))), mean_vega=('vega','mean'))
    cands['score'] = np.abs(cands['mean_abs_k'] - TARGET_ABS_K) - 0.05*np.log1p(cands['coverage']) - 0.01*np.log1p(cands['mean_vega'])
    chosen = cands.sort_values(['trade_date','score']).groupby('trade_date', as_index=False).first()[['trade_date','contract_id']]
    out = q.merge(chosen, on=['trade_date','contract_id'], how='inner').copy()
    out = out.sort_values(['trade_date','timestamp']).reset_index(drop=True)
    return out


def build_transitions(main_q: pd.DataFrame, feat: pd.DataFrame) -> pd.DataFrame:
    q = main_q.sort_values(['trade_date','contract_id','timestamp']).copy().reset_index(drop=True)
    same = q['trade_date'].eq(q['trade_date'].shift(-1)) & q['contract_id'].eq(q['contract_id'].shift(-1))
    one = (q['timestamp'].shift(-1) - q['timestamp']).dt.total_seconds().eq(60)
    t = q.loc[same & one].copy()
    for col in ['timestamp','mid','S','tau','iv']:
        t[f'{col}_next'] = q[col].shift(-1).loc[t.index].values
    t = t.merge(feat[['timestamp','trade_date','sigma_atm','skew','curvature_bfly','event_score','regime','regime_name']], on=['timestamp','trade_date'], how='inner')
    return t.reset_index(drop=True)


def run_policy(trans: pd.DataFrame, method: str, delta_kind: str, base_interval: int, fast_interval: int|None=None, stock_band: float|None=None):
    rows=[]
    held=0.0
    last_rehedge=None
    n_rehedges=0
    for row in trans.itertuples(index=False):
        if delta_kind=='bs':
            target = -bs_delta(row.cp, row.S, row.K, row.tau, row.iv)
        elif delta_kind=='surface':
            target = -surface_delta(row.cp, row.S, row.K, row.tau, row.sigma_atm, row.skew, row.curvature_bfly)
        elif delta_kind=='shock_surface':
            target = -shock_surface_delta(row.cp, row.S, row.K, row.tau, row.sigma_atm, row.skew, row.curvature_bfly, row.event_score, row.regime)
        else:
            raise ValueError(delta_kind)

        if last_rehedge is None:
            do=True
        else:
            mins = (row.timestamp - last_rehedge).total_seconds()/60.0
            interval = fast_interval if (fast_interval is not None and int(row.regime)==1) else base_interval
            do = mins >= interval
            if (not do) and (stock_band is not None) and abs(target-held) >= stock_band:
                do = True
        if do:
            held = target
            last_rehedge = row.timestamp
            n_rehedges += 1
        dS = float(row.S_next - row.S)
        dOpt = float(row.mid_next - row.mid)
        hedge_error = dOpt + held*dS
        rows.append((row.trade_date, row.timestamp, method, row.regime_name, int(row.regime), held, target, hedge_error, int(do)))
    pnl = pd.DataFrame(rows, columns=['trade_date','timestamp','method','regime_name','regime','held_stock','target_stock','hedge_error','rehedged'])
    overall = {
        'method': method,
        'mae': float(np.mean(np.abs(pnl['hedge_error']))),
        'rmse': float(np.sqrt(np.mean(np.square(pnl['hedge_error'])))),
        'p95_abs_err': float(np.quantile(np.abs(pnl['hedge_error']), 0.95)),
        'n_rehedges': int(pnl['rehedged'].sum()),
        'n_obs': int(len(pnl)),
        'stock_turnover': float(np.abs(pd.Series(pnl['held_stock']).diff().fillna(pnl['held_stock'])).sum()),
    }
    by_day = pnl.groupby('trade_date').agg(mae=('hedge_error', lambda s: float(np.mean(np.abs(s)))), rmse=('hedge_error', lambda s: float(np.sqrt(np.mean(np.square(s))))), n_rehedges=('rehedged','sum')).reset_index()
    by_day['method']=method
    by_regime = pnl.groupby(['regime_name']).agg(mae=('hedge_error', lambda s: float(np.mean(np.abs(s)))), n_obs=('hedge_error','size'), rehedges=('rehedged','sum')).reset_index()
    by_regime['rehedge_rate']=by_regime['rehedges']/by_regime['n_obs']
    by_regime['method']=method
    return pnl, overall, by_day, by_regime


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    all_qiv=[]
    all_feat=[]
    for exp in EXPIRIES:
        cs, ps = f'{exp} C', f'{exp} P'
        print('processing', exp, flush=True)
        if cs not in wb.sheetnames or ps not in wb.sheetnames:
            continue
        q = pd.concat([parse_sheet(wb[cs], cs), parse_sheet(wb[ps], ps)], ignore_index=True)
        prev_dates = sorted(pd.to_datetime(q.loc[q['trade_date'] < q['expiry_date'], 'trade_date']).dt.normalize().unique())
        if len(prev_dates) == 0:
            continue
        prev_date = pd.Timestamp(prev_dates[-1]).normalize()
        q = q[q['trade_date'].eq(prev_date)].copy()
        if q.empty:
            continue
        spot = robust_spot_from_pairs(q)
        if q.empty or spot.empty:
            continue
        qiv = iv_for_quotes(q, spot)
        if qiv.empty:
            continue
        feat = build_features(qiv)
        if feat.empty:
            continue
        all_qiv.append(qiv)
        all_feat.append(feat)
        # save a sample daily regime plot for the example date if present
    qiv = pd.concat(all_qiv, ignore_index=True).sort_values(['trade_date','timestamp'])
    feat = pd.concat(all_feat, ignore_index=True).sort_values(['trade_date','timestamp'])
    main_q = pick_main_contracts(qiv)
    trans = build_transitions(main_q, feat)
    qiv.to_csv(OUTDIR/'qiv_sample.csv', index=False)
    feat.to_csv(OUTDIR/'features_regimes_sample.csv', index=False)
    trans.to_csv(OUTDIR/'main_transitions_sample.csv', index=False)

    configs = [
        ('bs_1m','bs',1,None,None),
        ('surface_1m','surface',1,None,None),
        ('surface_5m','surface',5,None,None),
        ('surface_10m','surface',10,None,None),
        ('surface_regime','surface',10,1,None),
        ('surface_regime_band','surface',10,1,0.08),
        ('shock_surface_regime','shock_surface',10,1,None),
    ]
    panel_rows=[]; overall_rows=[]; by_day_rows=[]; by_regime_rows=[]
    for cfg in configs:
        pnl, overall, by_day, by_regime = run_policy(trans, *cfg)
        panel_rows.append(pnl); overall_rows.append(overall); by_day_rows.append(by_day); by_regime_rows.append(by_regime)
    panel = pd.concat(panel_rows, ignore_index=True)
    overall = pd.DataFrame(overall_rows).sort_values('mae').reset_index(drop=True)
    by_day = pd.concat(by_day_rows, ignore_index=True)
    by_regime = pd.concat(by_regime_rows, ignore_index=True)
    panel.to_csv(OUTDIR/'panel_results.csv', index=False)
    overall.to_csv(OUTDIR/'overall_summary.csv', index=False)
    by_day.to_csv(OUTDIR/'by_day_summary.csv', index=False)
    by_regime.to_csv(OUTDIR/'by_regime_summary.csv', index=False)

    # Graph 1 frontier
    plt.figure(figsize=(9,5.5))
    plt.scatter(overall['n_rehedges'], overall['mae'], s=100)
    for _, r in overall.iterrows():
        plt.annotate(r['method'], (r['n_rehedges'], r['mae']), xytext=(6,6), textcoords='offset points')
    plt.xlabel('Number of rehedges')
    plt.ylabel('MAE of hedge error')
    plt.title('Custom hedging experiment: MAE vs rehedge count')
    plt.tight_layout()
    plt.savefig(OUTDIR/'frontier_mae_vs_rehedges.png', dpi=180, bbox_inches='tight')
    plt.close()

    # Graph 2 regime rates
    methods = overall['method'].tolist()
    regimes = ['fast','slow'] if set(by_regime['regime_name']) >= {'fast','slow'} else sorted(by_regime['regime_name'].unique())
    x=np.arange(len(regimes)); width=0.11
    plt.figure(figsize=(11,5))
    for j,m in enumerate(methods):
        vals=[]
        for reg in regimes:
            sub=by_regime[(by_regime['method']==m)&(by_regime['regime_name']==reg)]
            vals.append(float(sub['rehedge_rate'].iloc[0]) if len(sub) else np.nan)
        plt.bar(x + (j-(len(methods)-1)/2)*width, vals, width=width, label=m)
    plt.xticks(x, regimes)
    plt.ylabel('Rehedge rate')
    plt.title('Rehedge rate by regime')
    plt.legend(fontsize=8, ncol=2)
    plt.tight_layout()
    plt.savefig(OUTDIR/'rehedge_rate_by_regime.png', dpi=180, bbox_inches='tight')
    plt.close()

    # Graph 3 normalized daily MAE vs bs_1m baseline
    base = by_day[by_day['method']=='bs_1m'][['trade_date','mae']].rename(columns={'mae':'base_mae'})
    merged = by_day.merge(base, on='trade_date', how='left')
    merged['mae_ratio_to_bs1m'] = merged['mae']/merged['base_mae']
    plt.figure(figsize=(10.5,5.5))
    for m in ['surface_1m','surface_5m','surface_10m','surface_regime','surface_regime_band','shock_surface_regime']:
        sub = merged[merged['method']==m].sort_values('trade_date')
        plt.plot(sub['trade_date'], sub['mae_ratio_to_bs1m'], marker='o', label=m)
    plt.axhline(1.0, linestyle='--')
    plt.ylabel('Daily MAE / bs_1m daily MAE')
    plt.xlabel('Trade date')
    plt.title('Daily MAE relative to Black--Scholes 1-minute hedge')
    plt.legend(fontsize=8, ncol=2)
    plt.tight_layout()
    plt.savefig(OUTDIR/'daily_mae_ratio_vs_bs1m.png', dpi=180, bbox_inches='tight')
    plt.close()

    # Graph 4 regime-specific MAE
    x=np.arange(len(regimes)); width=0.11
    plt.figure(figsize=(11,5))
    for j,m in enumerate(methods):
        vals=[]
        for reg in regimes:
            sub=by_regime[(by_regime['method']==m)&(by_regime['regime_name']==reg)]
            vals.append(float(sub['mae'].iloc[0]) if len(sub) else np.nan)
        plt.bar(x + (j-(len(methods)-1)/2)*width, vals, width=width, label=m)
    plt.xticks(x, regimes)
    plt.ylabel('MAE')
    plt.title('Hedge error by regime')
    plt.legend(fontsize=8, ncol=2)
    plt.tight_layout()
    plt.savefig(OUTDIR/'mae_by_regime.png', dpi=180, bbox_inches='tight')
    plt.close()

    summary = {
        'expiries_used': EXPIRIES,
        'n_trade_dates': int(trans['trade_date'].nunique()),
        'n_obs': int(len(trans)),
        'main_contracts_by_date': {
            str(k): v for k, v in main_q.groupby('trade_date')['contract_id'].first().to_dict().items()
        },
        'best_method_by_mae': overall.iloc[0].to_dict(),
    }
    with open(OUTDIR/'experiment_summary.json','w') as f:
        json.dump(summary, f, indent=2, default=str)
    print(overall.to_string(index=False))
    print('saved to', OUTDIR)

if __name__ == '__main__':
    main()
